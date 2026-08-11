"""
Génération de l'export Excel du calcul de CMPC (WACC), reproduisant la mise en
forme de la feuille "WACC - Indirect" du template DATA/wacc_template.xlsx, et
embarquant les feuilles Damodaran utilisées pour les calculs avec des liens
directs vers les lignes sourcées.
"""
import difflib
from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

DARK_BLUE = "3D4775"
RESULT_BLUE = "0071CE"
GREY_TEXT = "464B4B"
FOOTER_GREY = "787878"
LINK_BLUE = "0563C1"

FONT_NAME = "Calibri"
ROW_HEIGHT = 11

thin = Side(style="thin", color="000000")
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_TB = Border(top=thin, bottom=thin)

HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
HEADER_FONT = Font(name=FONT_NAME, size=8, bold=True, color="FFFFFF")

BOX_FILL = PatternFill("solid", fgColor=DARK_BLUE)
BOX_FONT = Font(name=FONT_NAME, size=8, bold=True, color="FFFFFF")

RESULT_FONT = Font(name=FONT_NAME, size=8, bold=True, color=RESULT_BLUE)

LABEL_FONT = Font(name=FONT_NAME, size=8, color=GREY_TEXT)
VALUE_FONT = Font(name=FONT_NAME, size=8, color=GREY_TEXT)
CODE_FONT = Font(name=FONT_NAME, size=8, color=GREY_TEXT)
SOURCE_FONT = Font(name=FONT_NAME, size=8, color=GREY_TEXT)
LINK_FONT = Font(name=FONT_NAME, size=8, color=LINK_BLUE, underline="single")
TITLE_FONT = Font(name=FONT_NAME, size=8, bold=True, color="FFFFFF")
META_FONT = Font(name=FONT_NAME, size=8, italic=True, color=GREY_TEXT)
FOOTER_FONT = Font(name=FONT_NAME, size=6, color=FOOTER_GREY)

PCT2 = "0.00%"
PCT0 = "0%"
NUM2 = "#,##0.00_);(#,##0.00);\"-\""
RATIO = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'

TREASURY_URL = "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_yield_curve"

DAMODARAN_SOURCES = [
    ("Betas par industrie (marchés émergents)", "betaemerg.xls", "Industry Averages",
     "https://pages.stern.nyu.edu/~adamodar/pc/datasets/betaemerg.xls",
     "Unlevered beta (colonne F) et D/E Ratio par secteur -> feuille 'Industry Averages'"),
    ("Primes de risque pays (ERP)", "ctryprem.xlsx", "ERPs by country",
     "https://www.stern.nyu.edu/~adamodar/pc/datasets/ctryprem.xlsx",
     "Country Risk Premium par pays et prime de risque marché mature (cellule E3) -> feuille 'ERPs by country'"),
    ("Taux d'imposition par pays", "countrytaxrates.xls", "feuille 1",
     "https://www.stern.nyu.edu/~adamodar/pc/datasets/countrytaxrates.xls",
     "Corporate tax rate par pays -> feuille 'Tax Rate'"),
    ("Spread de financement par industrie", "wacc.xls", "Industry Averages",
     "https://www.stern.nyu.edu/~adamodar/pc/datasets/wacc.xls",
     "Spread de dette et proportion E/(D+E) par secteur -> feuille 'Spread de financement US'"),
    ("Taux sans risque US", "Daily Treasury Par Yield Curve Rates", "BC_10YEAR / BC_30YEAR",
     TREASURY_URL,
     "Moyenne annuelle du taux des bons du Trésor US à 30 ans pour l'année de valorisation sélectionnée"),
]


def _find_match_row(df, value):
    """Reproduit la logique de matching (exact puis fuzzy) utilisée dans app.py
    et retourne l'index (position) de la ligne trouvée dans le DataFrame, ou None."""
    if df is None or len(df) == 0:
        return None
    col0 = df.columns[0]
    lower_series = df[col0].astype(str).str.strip().str.lower()
    target = str(value).strip().lower()

    exact = df[lower_series == target]
    if len(exact) > 0:
        return exact.index[0]

    candidates = lower_series.dropna().tolist()
    matches = difflib.get_close_matches(target, candidates, n=1, cutoff=0.9)
    if matches:
        match_rows = df[lower_series == matches[0]]
        if len(match_rows) > 0:
            return match_rows.index[0]

    return None


def _set_col_widths(ws):
    widths = {"A": 8.9, "B": 3.4, "C": 53.7, "D": 8.7, "E": 19.7, "F": 82.6, "G": 8.9}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _set_row_heights(ws, max_row):
    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT


def _title(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER_ALL


def _header_row(ws, row):
    labels = {3: "WACC", 4: "NEW", 5: "Formule", 6: "Commentaires/Sources"}
    for col in range(3, 7):
        cell = ws.cell(row=row, column=col, value=labels.get(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if col in (5, 6) else "left")
        cell.border = Border(left=thin, right=thin, top=thin)


def _line(ws, row, label, value, code, source=None, number_format=PCT2):
    c = ws.cell(row=row, column=3, value=label)
    c.font = LABEL_FONT
    d = ws.cell(row=row, column=4, value=value)
    d.font = VALUE_FONT
    d.number_format = number_format
    d.alignment = Alignment(horizontal="right")
    e = ws.cell(row=row, column=5, value=code)
    e.font = CODE_FONT
    if source is not None:
        f = ws.cell(row=row, column=6, value=source)
        f.font = SOURCE_FONT
    return ws.cell(row=row, column=6)


def _box_row(ws, row, label, value, code, number_format=PCT2):
    c = ws.cell(row=row, column=3, value=label)
    c.font = BOX_FONT
    c.fill = BOX_FILL
    c.border = BORDER_ALL
    d = ws.cell(row=row, column=4, value=value)
    d.font = BOX_FONT
    d.fill = BOX_FILL
    d.border = BORDER_ALL
    d.number_format = number_format
    d.alignment = Alignment(horizontal="right")
    e = ws.cell(row=row, column=5, value=code)
    e.font = BOX_FONT
    e.fill = BOX_FILL
    e.border = BORDER_ALL


def _result_row(ws, row, label, value, code, number_format=PCT2):
    c = ws.cell(row=row, column=3, value=label)
    c.font = RESULT_FONT
    c.border = BORDER_TB
    d = ws.cell(row=row, column=4, value=value)
    d.font = RESULT_FONT
    d.border = BORDER_TB
    d.number_format = number_format
    d.alignment = Alignment(horizontal="right")
    e = ws.cell(row=row, column=5, value=code)
    e.font = RESULT_FONT
    e.border = BORDER_TB


def _hyperlink_internal(cell, sheet_name, row, col=1):
    letter = get_column_letter(col)
    cell.hyperlink = f"#'{sheet_name}'!{letter}{row}"
    cell.font = LINK_FONT


def _hyperlink_external(cell, url):
    cell.hyperlink = url
    cell.font = LINK_FONT


def _write_dataframe_sheet(wb, sheet_name, df, prefix_rows=None):
    """Écrit un DataFrame brut (issu d'un fichier Damodaran) dans une nouvelle
    feuille du classeur. Retourne (worksheet, header_row) pour permettre le
    calcul du numéro de ligne Excel correspondant à un index de DataFrame."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    row = 1
    if prefix_rows:
        for label, value in prefix_rows:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            vcell = ws.cell(row=row, column=2, value=value)
            vcell.font = VALUE_FONT
            if isinstance(value, float):
                vcell.number_format = PCT2
            row += 1
        row += 1

    header_row = row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for offset, (_, data_row) in enumerate(df.iterrows(), start=1):
        excel_row = header_row + offset
        for col_idx, value in enumerate(data_row, start=1):
            if pd.isna(value):
                value = None
            ws.cell(row=excel_row, column=col_idx, value=value).font = VALUE_FONT

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        width = max(14, min(45, len(str(col_name)) + 6))
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    _set_row_heights(ws, header_row + len(df) + 1)

    return ws, header_row


def generate_wacc_workbook(data: dict, betas_data, erps_data, tax_rates_data, financing_spread_data) -> BytesIO:
    """
    Construit le classeur Excel d'export du calcul de CMPC : feuille "WACC" au
    même layout/mise en forme que le template, plus les feuilles de données
    Damodaran réellement utilisées, avec des liens directs entre les deux.

    `data` attend les clés :
    country, industry, valuation_year, avg_30y_rate (en %), country_risk_premium,
    beta_desendette, gearing_sectoriel, tax_rate, beta_reendette, c, d, e,
    cout_fonds_propres, adjusted_spread, is_adjusted, cout_dette,
    quote_part_equity, quote_part_debt, wacc, inflation_mature, inflation_locale,
    wacc_local
    """
    country = data["country"]
    industry = data["industry"]
    year = data["valuation_year"]

    wb = Workbook()
    ws = wb.active
    ws.title = "WACC"
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws)

    # Feuilles de données Damodaran embarquées
    industry_sheet, industry_header_row = _write_dataframe_sheet(wb, "Industry Averages", betas_data)
    tax_sheet, tax_header_row = _write_dataframe_sheet(wb, "Tax Rate", tax_rates_data)
    spread_sheet, spread_header_row = _write_dataframe_sheet(wb, "Spread de financement US", financing_spread_data)
    erps_sheet, erps_header_row = _write_dataframe_sheet(
        wb, "ERPs by country", erps_data,
        prefix_rows=[("Prime de risque marché mature (cellule E3 du fichier original ctryprem.xlsx)", data["c"])],
    )

    industry_idx = _find_match_row(betas_data, industry)
    tax_idx = _find_match_row(tax_rates_data, country)
    spread_idx = _find_match_row(financing_spread_data, industry)
    erps_idx = _find_match_row(erps_data, country)

    # ===== Feuille WACC =====
    _title(ws, 1, "Méthode Indirecte")

    meta = ws.cell(
        row=3, column=3,
        value=f"Pays : {country}    |    Industrie : {industry}    |    Année de valorisation : {year}"
              f"    |    Généré le : {datetime.now().strftime('%d/%m/%Y')}",
    )
    meta.font = META_FONT

    _header_row(ws, 5)

    maturity = data.get("risk_free_maturity", "30Y")
    f6 = _line(ws, 6, "Taux sans risque", data["avg_30y_rate"] / 100, "E",
               f"US Treasury {maturity} Bond rate - moyenne {year} (home.treasury.gov)")
    _hyperlink_external(f6, TREASURY_URL)

    f7 = _line(ws, 7, "Prime de risque pays", data["country_risk_premium"], "G",
               f"Prime de risque pays {country} - Damodaran au 01/01/{year} (ctryprem.xlsx, ERPs by country)")
    if erps_idx is not None:
        _hyperlink_internal(f7, "ERPs by country", erps_header_row + 1 + erps_idx)

    _line(ws, 8, "Taux sans risque local", "=SUM(D6:D7)", "I=E+G")

    f9 = _line(ws, 9, "Beta désendetté", data["beta_desendette"], "A",
               f"Beta pays émergent {industry} - Damodaran au 01/01/{year} (betaemerg.xls, Industry Averages, colonne F)",
               number_format=NUM2)
    if industry_idx is not None:
        _hyperlink_internal(f9, "Industry Averages", industry_header_row + 1 + industry_idx)

    f10 = _line(ws, 10, "Gearing : dette nette / fonds propres", data["gearing_sectoriel"], "B",
                f"Gearing {industry} - Damodaran au 01/01/{year} (betaemerg.xls, Industry Averages, colonne D/E Ratio)")
    if industry_idx is not None:
        _hyperlink_internal(f10, "Industry Averages", industry_header_row + 1 + industry_idx)

    f11 = _line(ws, 11, "Taux d'impôt", data["tax_rate"], "C",
                f"Taux d'IS {country} - Damodaran (countrytaxrates.xls)", number_format=PCT0)
    if tax_idx is not None:
        _hyperlink_internal(f11, "Tax Rate", tax_header_row + 1 + tax_idx)

    _line(ws, 12, "Beta réendetté", "=D9*(1+D10*(1-D11))", "D=A*(1+B*(1-C))", number_format=NUM2)

    f13 = _line(ws, 13, "Prime de risque du marché actions", data["c"], "F",
                f"Prime de risque des marchés matures - Damodaran au 01/01/{year} (ctryprem.xlsx, ERPs by country, cellule E3)")
    _hyperlink_internal(f13, "ERPs by country", 1, col=2)

    _line(ws, 14, "Prime de taille", data["d"] + data["e"], "H",
          "Prime de taille + prime spécifique - Saisie manuelle")

    _box_row(ws, 16, "Coût des fonds propres", '=IFERROR(D12*D13+D8+D14,"na")', "J=I+D*F+H")

    _line(ws, 18, "Taux sans risque", "=D6", "E", "=F6")

    f19 = _line(ws, 19, "Prime de risque pays", "=D7", "G",
                f"Damodaran - Spread par défaut basé sur la notation de {country}")
    if erps_idx is not None:
        _hyperlink_internal(f19, "ERPs by country", erps_header_row + 1 + erps_idx)

    _line(ws, 20, "Taux sans risque local", "=SUM(D18:D19)", "I=E+G")

    spread_source = "Damodaran - Spread de la dette basé sur le secteur d'activité, US (wacc.xls, Industry Averages)"
    if data.get("is_adjusted"):
        spread_source += " - ajusté (table de correction du spread, faible cost of debt)"
    f21 = _line(ws, 21, "Spread de dette", data["adjusted_spread"], "L", spread_source)
    if spread_idx is not None:
        _hyperlink_internal(f21, "Spread de financement US", spread_header_row + 1 + spread_idx)

    _line(ws, 22, "Taux d'impôt", "=D11", "C", "=F11")

    _box_row(ws, 24, "Coût de la dette après impôts", '=IFERROR((D20+D21)*(1-D22),"na")', "N=(I+L)*(1-C)",
              number_format="0.0%")

    _line(ws, 26, "Fonds Propres / Valeur d'Entp.", "=1/(1+D10)", "P=1/(1+B)", number_format=RATIO)
    _line(ws, 27, "Dette nette / Valeur d'Entp.", "=1-D26", "O=B/(1+B)", number_format=RATIO)

    _result_row(ws, 29, "CMPC - monnaie du taux sans risque", '=IFERROR(D16*D26+D24*D27,"na")', "Q=J*P+N*O")

    _line(ws, 31, "Inflation, monnaie du taux sans risque", data["inflation_mature"], "R",
          "Saisie manuelle (ex. Cleveland Fed - Ten year expected inflation)", number_format="0.0%")
    _line(ws, 32, "Inflation, monnaie locale", data["inflation_locale"], "S",
          "Saisie manuelle (ex. critère de convergence UEMOA, 3% plafond)", number_format="0.0%")

    _result_row(ws, 34, "CMPC en monnaie locale", "=(1+D29)*(1+D32)/(1+D31)-1", "T=(1+J)*(1+S)/(1+R)-1")

    footer = ws.cell(row=36, column=3, value="Source: Damodaran (NYU Stern) & US Treasury")
    footer.font = FOOTER_FONT

    _set_row_heights(ws, 36)

    _add_sources_sheet(wb)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _add_sources_sheet(wb: Workbook):
    ws = wb.create_sheet("Sources Damodaran")
    ws.sheet_view.showGridLines = False
    widths = {"A": 32, "B": 24, "C": 20, "D": 60, "E": 65}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    headers = ["Donnée", "Fichier source", "Feuille (originale)", "Lien", "Utilisation dans le calcul"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, (name, filename, sheet, url, usage) in enumerate(DAMODARAN_SOURCES, start=2):
        ws.cell(row=i, column=1, value=name).font = LABEL_FONT
        ws.cell(row=i, column=2, value=filename).font = VALUE_FONT
        ws.cell(row=i, column=3, value=sheet).font = VALUE_FONT
        link_cell = ws.cell(row=i, column=4, value=url)
        _hyperlink_external(link_cell, url)
        ws.cell(row=i, column=5, value=usage).font = SOURCE_FONT

    _set_row_heights(ws, len(DAMODARAN_SOURCES) + 1)
