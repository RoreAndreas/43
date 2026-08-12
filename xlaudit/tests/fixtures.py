"""Construction des classeurs de fixtures.

Deux classeurs de meme mise en page sont produits : un sain et un porteur d'une
anomalie de chaque type recherche. La comparaison des deux donne son sens au test
de non-detection, qui compte autant que le test de detection.

Point technique : openpyxl ecrit des formules **sans valeur en cache**. Un
classeur de fixture relu avec `data_only=True` rendrait donc `None` partout, et
il serait impossible de tester R201 (erreurs en cache) ou le chiffreur d'ecart.
Les valeurs sont donc injectees dans le XML apres coup, exactement comme Excel les
y aurait ecrites. Les valeurs injectees sont celles qu'Excel calculerait avec la
formule *reellement presente* : c'est ce qui rend verifiable le test d'egalite
exacte de R205.
"""

from __future__ import annotations

import shutil
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

from xlaudit.core.refs import index_to_col

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ERROR_STRINGS = {"#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"}


class FixtureBuilder:
    """Declare des cellules (formule et/ou valeur) puis ecrit le classeur."""

    def __init__(self) -> None:
        self.sheets: dict[str, dict[tuple[int, int], tuple[str | None, object]]] = {}

    def cell(self, sheet: str, row: int, col: int, formula: str | None = None, value=None):
        self.sheets.setdefault(sheet, {})[(row, col)] = (formula, value)
        return self

    def text(self, sheet: str, row: int, col: int, value: str):
        return self.cell(sheet, row, col, None, value)

    def save(self, path: str | Path) -> Path:
        path = Path(path)
        wb = openpyxl.Workbook()
        default = wb.active
        first = True
        for name, cells in self.sheets.items():
            if first:
                ws = default
                ws.title = name
                first = False
            else:
                ws = wb.create_sheet(name)
            for (row, col), (formula, value) in cells.items():
                ws.cell(row=row, column=col).value = formula if formula else value
        wb.save(path)
        wb.close()

        cached = {
            name: {rc: val for rc, (f, val) in cells.items() if f is not None}
            for name, cells in self.sheets.items()
        }
        inject_cached_values(path, cached)
        return path


def _sheet_xml_map(zf: zipfile.ZipFile) -> dict[str, str]:
    """Nom d'onglet -> chemin du XML dans l'archive."""
    root = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    target_by_id = {
        r.get("Id"): r.get("Target", "") for r in rels.findall(f"{{{NS_PKG_REL}}}Relationship")
    }
    out: dict[str, str] = {}
    for sh in root.findall(f"{{{NS_MAIN}}}sheets/{{{NS_MAIN}}}sheet"):
        rid = sh.get(f"{{{NS_REL}}}id")
        target = target_by_id.get(rid, "")
        if not target:
            continue
        if target.startswith("/"):
            target = target.lstrip("/")
        elif not target.startswith("xl/"):
            target = "xl/" + target
        out[sh.get("name", "")] = target
    return out


def inject_cached_values(path: str | Path, values: dict[str, dict[tuple[int, int], object]]) -> None:
    """Ecrit une valeur en cache `<v>` a cote de chaque `<f>`, comme le ferait Excel.

    Les erreurs prennent `t="e"`, les chaines `t="str"`, les nombres restent au
    type numerique par defaut.
    """
    path = Path(path)
    tmp = path.with_suffix(path.suffix + ".tmp")
    ET.register_namespace("", NS_MAIN)

    with zipfile.ZipFile(path) as zin:
        sheet_map = _sheet_xml_map(zin)
        entries = {n: zin.read(n) for n in zin.namelist()}

    for sheet_name, cell_values in values.items():
        xml_path = sheet_map.get(sheet_name)
        if not xml_path or xml_path not in entries:
            continue
        by_ref = {
            f"{index_to_col(c)}{r}": v for (r, c), v in cell_values.items() if v is not None
        }
        if not by_ref:
            continue
        root = ET.fromstring(entries[xml_path])
        for cell in root.iter(f"{{{NS_MAIN}}}c"):
            ref = cell.get("r")
            if ref not in by_ref:
                continue
            if cell.find(f"{{{NS_MAIN}}}f") is None:
                continue
            for old in cell.findall(f"{{{NS_MAIN}}}v"):
                cell.remove(old)
            value = by_ref[ref]
            node = ET.SubElement(cell, f"{{{NS_MAIN}}}v")
            if isinstance(value, str) and value in ERROR_STRINGS:
                cell.set("t", "e")
                node.text = value
            elif isinstance(value, str):
                cell.set("t", "str")
                node.text = value
            elif isinstance(value, bool):
                cell.set("t", "b")
                node.text = "1" if value else "0"
            else:
                if "t" in cell.attrib:
                    del cell.attrib["t"]
                node.text = repr(float(value)) if isinstance(value, float) else str(value)
        entries[xml_path] = ET.tostring(root, encoding="UTF-8", xml_declaration=True)

    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)
    shutil.move(str(tmp), str(path))


# ---------------------------------------------------------------------------
# Le modele de fixture
# ---------------------------------------------------------------------------
# Mise en page (onglet "Modele") :
#   colonne A  libelles
#   colonnes D..I (4..9)  six periodes de projection
#   lignes  6..9    produits (constantes d'entree)
#   ligne   10      total produits
#   lignes  12..15  charges (constantes d'entree)
#   ligne   16      total charges
#   ligne   18      marge
#   ligne   20      marge nette          <- R202 (valeur en dur)
#   ligne   22      indice               <- R203 (variante horizontale)
#   lignes  30..33  recap                <- R204 (glissement de recopie)
#   ligne   35      produits cumules     <- R205 (ancrage asymetrique)
#   ligne   37      total charges bis    <- R206 (somme tronquee)
#   ligne   39      somme par categorie  <- R207 (plage de somme multi-colonnes)
#   ligne   41      total general        <- R208 (somme heterogene)
#   lignes  44..47  petite table de categories
#   ligne   49      impasse de tresorerie<- R215 (ligne orpheline significative)
#   ligne   53      synthese (consommateur de la ligne 49 dans le classeur sain)
#   ligne   55      controle equilibre   <- R202 haute confiance
#   ligne   56      service de la dette
#   ligne   57      ratio de couverture  <- R201 (#DIV/0!)
#   ligne   59      poste supprime       <- R201 (#REF!)

SHEET = "Modele"
TIES_SHEET = "Etats fi periode"
FIRST_COL = 4  # D
N_PER = 6
COLS = list(range(FIRST_COL, FIRST_COL + N_PER))


def _produit(k: int, p: int) -> float:
    return float(100 * (k + 1) + 10 * p)


def _charge(k: int, p: int) -> float:
    return float(40 * (k + 1) + 3 * p)


def _total_produits(p: int) -> float:
    return sum(_produit(k, p) for k in range(4))


def _total_charges(p: int) -> float:
    return sum(_charge(k, p) for k in range(4))


def _marge(p: int) -> float:
    return _total_produits(p) - _total_charges(p)


def build_model(path: str | Path, faulty: bool) -> Path:
    """Construit le classeur de fixture, sain ou porteur des anomalies."""
    b = FixtureBuilder()
    S = SHEET

    def col(p: int) -> int:
        return FIRST_COL + p

    def letter(p: int) -> str:
        return index_to_col(col(p))

    # -- en-tetes --------------------------------------------------------
    b.text(S, 1, 1, "Modele de projection - fixture xlaudit")
    b.text(S, 3, 1, "Periode")
    for p in range(N_PER):
        b.cell(S, 3, col(p), None, p + 1)
        b.cell(S, 4, col(p), None, f"P{p + 1}")

    # -- produits (entrees) ----------------------------------------------
    b.text(S, 5, 1, "Produits")
    for k in range(4):
        b.text(S, 6 + k, 1, f"Produit {chr(65 + k)}")
        for p in range(N_PER):
            b.cell(S, 6 + k, col(p), None, _produit(k, p))
    b.text(S, 10, 1, "Total produits")
    for p in range(N_PER):
        b.cell(S, 10, col(p), f"=SUM({letter(p)}6:{letter(p)}9)", _total_produits(p))

    # -- charges (entrees) -----------------------------------------------
    b.text(S, 11, 1, "Charges")
    for k in range(4):
        b.text(S, 12 + k, 1, f"Charge {chr(65 + k)}")
        for p in range(N_PER):
            b.cell(S, 12 + k, col(p), None, _charge(k, p))
    b.text(S, 16, 1, "Total charges")
    for p in range(N_PER):
        b.cell(S, 16, col(p), f"=SUM({letter(p)}12:{letter(p)}15)", _total_charges(p))

    # -- marge ------------------------------------------------------------
    b.text(S, 18, 1, "Marge")
    for p in range(N_PER):
        b.cell(S, 18, col(p), f"={letter(p)}10-{letter(p)}16", _marge(p))

    # -- R202 : valeur en dur dans une ligne de formules homogene ---------
    b.text(S, 20, 1, "Marge nette")
    for p in range(N_PER):
        val = _marge(p) * 0.7
        if faulty and p == 2:
            b.cell(S, 20, col(p), None, val)  # constante isolee
        else:
            b.cell(S, 20, col(p), f"={letter(p)}18*0.7", val)

    # -- R203 : variante structurelle minoritaire -------------------------
    b.text(S, 22, 1, "Indice de marge")
    for p in range(N_PER):
        if faulty and p == 3:
            # Denominateur relatif au lieu d'ancre : meme resultat ici, structure differente.
            b.cell(S, 22, col(p), f"={letter(p)}18/{letter(0)}18", _marge(p) / _marge(0))
        else:
            b.cell(S, 22, col(p), f"={letter(p)}18/$D$18", _marge(p) / _marge(0))

    # -- R204 : glissement de recopie --------------------------------------
    b.text(S, 29, 1, "Recapitulatif produits")
    for k in range(4):
        row = 30 + k
        b.text(S, row, 1, f"Recap {chr(65 + k)}")
        for p in range(N_PER):
            if faulty:
                lo, hi = 6 + k, 9 + k
                formula = f"=SUM({letter(p)}{lo}:{letter(p)}{hi})"
                value = _sum_rows(lo, hi, p)
            else:
                formula = f"=SUM({letter(p)}$6:{letter(p)}$9)"
                value = _total_produits(p)
            b.cell(S, row, col(p), formula, value)

    # -- R205 : ancrage asymetrique ----------------------------------------
    b.text(S, 35, 1, "Produits (report)")
    for p in range(N_PER):
        if faulty:
            # $D6:<col>9 : le coin initial est ancre en colonne, le coin final ne
            # l'est pas. Recopiee vers la droite, la plage s'elargit au lieu de
            # se deplacer, et le terme cumule les periodes precedentes.
            formula = f"=SUM($D6:{letter(p)}9)"
            value = sum(_total_produits(q) for q in range(p + 1))
        else:
            formula = f"=SUM({letter(p)}6:{letter(p)}9)"
            value = _total_produits(p)
        b.cell(S, 35, col(p), formula, value)

    # -- R206 : somme tronquee ---------------------------------------------
    b.text(S, 37, 1, "Total charges (report)")
    for p in range(N_PER):
        if faulty and p == 2:
            formula = f"=SUM({letter(p)}12:{letter(p)}14)"  # omet la 4e ligne du bloc
            value = sum(_charge(k, p) for k in range(3))
        else:
            formula = f"=SUM({letter(p)}12:{letter(p)}15)"
            value = _total_charges(p)
        b.cell(S, 37, col(p), formula, value)

    # -- petite table de categories (support de R207) ----------------------
    keys = ["A", "B", "A", "B"]
    vals_e = [10.0, 20.0, 30.0, 40.0]
    vals_f = [1.0, 2.0, 3.0, 4.0]
    b.text(S, 43, 1, "Table de categories")
    for i in range(4):
        b.cell(S, 44 + i, 4, None, keys[i])
        b.cell(S, 44 + i, 5, None, vals_e[i])
        b.cell(S, 44 + i, 6, None, vals_f[i])

    # -- R207 : plage de somme plus large que la plage de criteres ---------
    b.text(S, 39, 1, "Somme categorie A")
    somme_e = sum(v for k, v in zip(keys, vals_e) if k == "A")
    if faulty:
        # Excel tronque la plage de somme a la largeur de la plage de criteres :
        # seule la colonne E est sommee, alors que la formule en cite deux.
        b.cell(S, 39, 4, '=SUMIF($D$44:$D$47,"A",$E$44:$F$47)', somme_e)
    else:
        b.cell(S, 39, 4, '=SUMIF($D$44:$D$47,"A",$E$44:$E$47)', somme_e)

    # -- R208 : nombre de termes variable selon la colonne -----------------
    b.text(S, 41, 1, "Total general")
    for p in range(N_PER):
        if faulty and p == 2:
            formula = f"={letter(p)}10+{letter(p)}16+{letter(p)}18"
            value = _total_produits(p) + _total_charges(p) + _marge(p)
        else:
            formula = f"={letter(p)}10+{letter(p)}16"
            value = _total_produits(p) + _total_charges(p)
        b.cell(S, 41, col(p), formula, value)

    # -- R215 : ligne orpheline a libelle significatif ---------------------
    b.text(S, 49, 1, "Impasse de tresorerie")
    for p in range(N_PER):
        b.cell(S, 49, col(p), f"={letter(p)}18-{letter(p)}16", _marge(p) - _total_charges(p))
    if not faulty:
        # Dans le classeur sain, l'impasse est consommee : le canal est branche.
        b.text(S, 53, 1, "Synthese")
        b.cell(S, 53, 4, "=SUM(D49:I49)", sum(_marge(p) - _total_charges(p) for p in range(N_PER)))

    # -- R202 haute confiance : ligne de controle --------------------------
    b.text(S, 55, 1, "Controle equilibre")
    for p in range(N_PER):
        if faulty and p == 4:
            b.cell(S, 55, col(p), None, 0.0)  # controle neutralise en dur
        else:
            b.cell(
                S, 55, col(p),
                f"=ROUND({letter(p)}18-({letter(p)}10-{letter(p)}16),8)", 0.0,
            )

    # -- R201 : erreurs en cache -------------------------------------------
    b.text(S, 56, 1, "Service de la dette")
    for p in range(N_PER):
        val = 0.0 if (faulty and p == 2) else float(100 + 10 * p)
        b.cell(S, 56, col(p), None, val)
    b.text(S, 57, 1, "Ratio de couverture")
    for p in range(N_PER):
        if faulty and p == 2:
            b.cell(S, 57, col(p), f"={letter(p)}18/{letter(p)}56", "#DIV/0!")
        else:
            b.cell(S, 57, col(p), f"={letter(p)}18/{letter(p)}56", _marge(p) / (100 + 10 * p))
    if faulty:
        b.text(S, 59, 1, "Poste supprime")
        b.cell(S, 59, 4, "=#REF!*2", "#REF!")

    _build_ties_sheet(b, faulty)
    return b.save(path)


def _sum_rows(lo: int, hi: int, p: int) -> float:
    """Valeur d'un SUM vertical sur l'onglet Modele, colonne de la periode p.

    Reproduit ce qu'Excel calculerait : les lignes vides valent zero, la ligne 10
    porte le total produits, la ligne 16 le total charges.
    """
    total = 0.0
    for r in range(lo, hi + 1):
        if 6 <= r <= 9:
            total += _produit(r - 6, p)
        elif r == 10:
            total += _total_produits(p)
        elif 12 <= r <= 15:
            total += _charge(r - 12, p)
        elif r == 16:
            total += _total_charges(p)
    return total


def _build_ties_sheet(b: FixtureBuilder, faulty: bool) -> None:
    """Onglet d'etats financiers, support des bouclages de phase 3.

    Le desequilibre plante vaut un centime sur une seule periode : c'est l'ordre
    de grandeur qui echappe a la relecture visuelle.
    """
    S = TIES_SHEET
    first = 15  # colonne O
    n = 6

    def col(p: int) -> int:
        return first + p

    b.text(S, 1, 1, "Etats financiers par periode - fixture")
    b.text(S, 9, 1, "Date")
    b.text(S, 10, 1, "Periode")
    for p in range(n):
        b.cell(S, 9, col(p), None, f"31/12/{2025 + p}")
        b.cell(S, 10, col(p), None, p + 1)

    b.text(S, 15, 1, "Tresorerie")
    b.text(S, 20, 1, "Total actif")
    b.text(S, 30, 1, "Total passif")
    b.text(S, 40, 1, "Solde de cloture")
    b.text(S, 41, 1, "Solde d'ouverture")

    for p in range(n):
        c = col(p)
        letter = index_to_col(c)
        tresorerie = 1000.0 + 100.0 * p
        actif = 5000.0 + 250.0 * p
        passif = actif + (0.01 if (faulty and p == 3) else 0.0)
        b.cell(S, 15, c, None, tresorerie)
        b.cell(S, 20, c, f"=3000+2000+{letter}15-{letter}15+{2000 + 250 * p}", actif)
        b.cell(S, 30, c, None, passif)
        # Cascade de tresorerie : le solde de cloture doit egaler la tresorerie du bilan.
        b.cell(S, 40, c, f"={letter}15", tresorerie)
        if p == 0:
            b.cell(S, 41, c, None, 0.0)
        else:
            prev = index_to_col(col(p - 1))
            b.cell(S, 41, c, f"={prev}40", 1000.0 + 100.0 * (p - 1))


def build_healthy(path: str | Path) -> Path:
    return build_model(path, faulty=False)


def build_faulty(path: str | Path) -> Path:
    return build_model(path, faulty=True)
