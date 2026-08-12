"""Tests du socle de chargement.

L'iteration sure est le point le plus dangereux du projet : en lecture seule, une
cellule vide n'a ni `.row` ni `.column`, et une iteration sans bornes explicites
ne commence pas la ou `ws.min_row` le laisse croire. Un bug ici contaminerait
toutes les regles, donc ces cas sont testes en premier.
"""

from __future__ import annotations

import openpyxl
import pytest
from openpyxl.worksheet.formula import ArrayFormula

from xlaudit.core import loader as L

from .fixtures import build_faulty, build_healthy


@pytest.fixture(scope="module")
def healthy(tmp_path_factory):
    return build_healthy(tmp_path_factory.mktemp("wb") / "sain.xlsx")


@pytest.fixture(scope="module")
def faulty(tmp_path_factory):
    return build_faulty(tmp_path_factory.mktemp("wb") / "anomalies.xlsx")


class TestSafeIteration:
    """Le piege des EmptyCell et de l'origine d'iteration."""

    @pytest.fixture
    def sparse(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        # Trous volontaires : lignes 3, 4 et 6 entierement vides, colonne A vide.
        ws["B2"] = 1
        ws["D2"] = "=SUM(B2:C2)"
        ws["B5"] = "libelle"
        ws["F7"] = 3
        path = tmp_path / "sparse.xlsx"
        wb.save(path)
        wb.close()
        return path

    def test_coordinates_are_exact(self, sparse):
        wb = openpyxl.load_workbook(sparse, read_only=True, data_only=False)
        try:
            got = {(r, c): v for r, c, v in L.iter_cells(wb["S"])}
        finally:
            wb.close()
        assert got == {
            (2, 2): 1,
            (2, 4): "=SUM(B2:C2)",
            (5, 2): "libelle",
            (7, 6): 3,
        }

    def test_does_not_touch_row_attribute_on_empty_cells(self, sparse):
        """Une EmptyCell n'a pas d'attribut .row : le prouver explicitement."""
        wb = openpyxl.load_workbook(sparse, read_only=True, data_only=False)
        try:
            ws = wb["S"]
            empties = [
                cell
                for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=6)
                for cell in row
                if type(cell).__name__ == "EmptyCell"
            ]
            assert empties, "la fixture doit contenir des cellules vides"
            for cell in empties:
                assert not hasattr(cell, "row")
                assert not hasattr(cell, "column")
            # ... et pourtant l'iteration sure fonctionne sur le meme onglet.
            assert len(list(L.iter_cells(ws))) == 4
        finally:
            wb.close()

    def test_bounded_iteration_offsets_columns_correctly(self, sparse):
        wb = openpyxl.load_workbook(sparse, read_only=True, data_only=False)
        try:
            got = list(L.iter_cells(wb["S"], min_row=2, max_row=2, min_col=3, max_col=6))
        finally:
            wb.close()
        assert got == [(2, 4, "=SUM(B2:C2)")]

    def test_empty_sheet_yields_nothing(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "vide"
        path = tmp_path / "vide.xlsx"
        wb.save(path)
        wb.close()
        wb2 = openpyxl.load_workbook(path, read_only=True)
        try:
            assert list(L.iter_cells(wb2["vide"])) == []
        finally:
            wb2.close()

    def test_sheet_without_dimension_element_is_still_read(self, tmp_path):
        """L'element <dimension> est facultatif et souvent absent.

        Sans lui, `ws.max_row` vaut None en lecture seule, et un scan naif lit
        zero cellule *sans lever d'erreur* : il reussit et ne trouve rien. C'est
        le pire mode de defaillance possible pour un outil d'audit.
        """
        import openpyxl as _op

        wb = _op.Workbook(write_only=True)  # n'ecrit pas <dimension>
        ws = wb.create_sheet("Sans dimension")
        ws.append(["Poste", 10, 20])
        ws.append(["Total", "=B1+C1", None])
        path = tmp_path / "sans_dimension.xlsx"
        wb.save(path)
        wb.close()

        import zipfile

        with zipfile.ZipFile(path) as zf:
            xml = zf.read("xl/worksheets/sheet1.xml").decode()
        assert "<dimension" not in xml, "la fixture doit vraiment omettre la dimension"

        wb2 = _op.load_workbook(path, read_only=True, data_only=False)
        try:
            assert wb2["Sans dimension"].max_row is None
        finally:
            wb2.close()

        snap = L.collect(path)
        assert snap.formula_count == 1
        assert snap.formula("Sans dimension", 2, 2) == "=B1+C1"
        assert snap.value("Sans dimension", 1, 2) == 10

    def test_array_formula_is_normalized_to_text(self, tmp_path):
        """Sans normalisation, les formules matricielles disparaissent des scans."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 1
        ws["C3"] = ArrayFormula("C3", "=SUM(A1:A5*2)")
        path = tmp_path / "array.xlsx"
        wb.save(path)
        wb.close()

        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=False)
        try:
            cells = {(r, c): v for r, c, v in L.iter_cells(wb2["S"])}
        finally:
            wb2.close()
        assert cells[(3, 3)] == "=SUM(A1:A5*2)"
        assert isinstance(cells[(3, 3)], str)


class TestFixtureIntegrity:
    """La fixture doit vraiment porter des valeurs en cache, sinon rien n'est testable."""

    def test_cached_values_are_present(self, healthy):
        wb = openpyxl.load_workbook(healthy, read_only=True, data_only=True)
        try:
            ws = wb["Modele"]
            values = {(r, c): v for r, c, v in L.iter_cells(ws)}
        finally:
            wb.close()
        # D10 = SUM(D6:D9) = 100+200+300+400
        assert values[(10, 4)] == pytest.approx(1000.0)

    def test_error_values_survive_the_roundtrip(self, faulty):
        wb = openpyxl.load_workbook(faulty, read_only=True, data_only=True)
        try:
            values = {(r, c): v for r, c, v in L.iter_cells(wb["Modele"])}
        finally:
            wb.close()
        assert values[(57, 6)] == "#DIV/0!"
        assert values[(59, 4)] == "#REF!"

    def test_formulas_and_values_come_from_different_loads(self, healthy):
        snap = L.collect(healthy)
        assert snap.formula("Modele", 10, 4) == "=SUM(D6:D9)"
        assert snap.value("Modele", 10, 4) == pytest.approx(1000.0)


class TestCollect:
    def test_sheets_and_counts(self, healthy):
        snap = L.collect(healthy)
        assert "Modele" in snap.sheet_names
        assert "Etats fi periode" in snap.sheet_names
        assert snap.formula_count > 50
        assert snap.stats["has_cached_values"] is True

    def test_sheet_lookup_is_case_insensitive(self, healthy):
        snap = L.collect(healthy)
        assert snap.sheet("modele") is not None
        assert snap.sheet("modele").name == "Modele"

    def test_row_labels(self, healthy):
        snap = L.collect(healthy)
        assert snap.row_label("Modele", 10) == "Total produits"
        assert snap.row_label("Modele", 49) == "Impasse de tresorerie"

    def test_label_zone_excludes_projection_columns(self, healthy):
        snap = L.collect(healthy)
        # Les formules commencent en colonne D : la zone de libelles s'arrete avant.
        assert snap.sheet("Modele").label_max_col < 4

    def test_source_file_is_never_modified(self, tmp_path):
        path = build_healthy(tmp_path / "intact.xlsx")
        before = L.sha256_file(path)
        snap = L.collect(path)
        _ = snap.formula_count
        assert L.sha256_file(path) == before


class TestCalcProperties:
    def test_calc_properties_are_read(self, healthy):
        # openpyxl ecrit calcPr ; on verifie surtout que la lecture ne casse pas.
        props = L.read_calc_properties(healthy)
        assert isinstance(props, dict)

    def test_missing_file_is_tolerated(self, tmp_path):
        assert L.read_calc_properties(tmp_path / "absent.xlsx") == {}


class TestDefinedNames:
    def test_sheet_scoped_names_are_visible(self, tmp_path):
        """`wb.defined_names` ne montre que la portee globale : on lit le XML."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feuille1"
        ws["A1"] = 1
        from openpyxl.workbook.defined_name import DefinedName

        wb.defined_names.add(DefinedName("GlobalName", attr_text="Feuille1!$A$1"))
        ws.defined_names.add(DefinedName("LocalName", attr_text="Feuille1!$A$1"))
        path = tmp_path / "names.xlsx"
        wb.save(path)
        wb.close()

        names = L.read_defined_names_xml(path)
        by_name = {n.name: n for n in names}
        assert "GlobalName" in by_name and by_name["GlobalName"].scope is None
        assert "LocalName" in by_name, "le nom a portee feuille doit etre visible"
        assert by_name["LocalName"].scope == "Feuille1"


class TestCache:
    def test_second_call_hits_the_cache(self, tmp_path):
        path = build_healthy(tmp_path / "cache.xlsx")
        cache_dir = tmp_path / "cache"
        snap1, hit1 = L.get_snapshot(path, cache_dir=cache_dir)
        snap2, hit2 = L.get_snapshot(path, cache_dir=cache_dir)
        assert hit1 is False and hit2 is True
        assert snap1.formula_count == snap2.formula_count
        assert snap2.formula("Modele", 10, 4) == "=SUM(D6:D9)"

    def test_cache_is_invalidated_by_content_change(self, tmp_path):
        path = tmp_path / "evolutif.xlsx"
        build_healthy(path)
        cache_dir = tmp_path / "cache"
        snap1, _ = L.get_snapshot(path, cache_dir=cache_dir)
        build_faulty(path)  # meme nom, contenu different
        snap2, hit = L.get_snapshot(path, cache_dir=cache_dir)
        assert hit is False, "un contenu different ne doit jamais etre servi par le cache"
        assert snap1.sha256 != snap2.sha256

    def test_roundtrip_preserves_values_and_labels(self, tmp_path):
        path = build_faulty(tmp_path / "rt.xlsx")
        snap = L.collect(path)
        out = tmp_path / "snap.json.gz"
        snap.save(out)
        back = L.Snapshot.load(out)
        assert back.sha256 == snap.sha256
        assert back.formula("Modele", 35, 5) == snap.formula("Modele", 35, 5)
        assert back.value("Modele", 57, 6) == "#DIV/0!"
        assert back.row_label("Modele", 49) == "Impasse de tresorerie"
