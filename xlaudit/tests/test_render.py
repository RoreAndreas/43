"""Tests des rendus.

Le point critique est le classeur de constats : il cite des formules a chaque
ligne, et une chaine commencant par `=` ecrite sans precaution redevient une
formule vivante a l'ouverture d'Excel.
"""

from __future__ import annotations

import json
import re
import zipfile

import openpyxl
import pytest

from xlaudit.core.graph import build_graph
from xlaudit.models import Reserve, RunLog, RuleRun, Signal, signals_to_document
from xlaudit.render.graphviz import sheet_map_dot, sheet_map_mermaid, trace_mermaid
from xlaudit.render.markdown import render_markdown
from xlaudit.render.xlsx import write_report, write_text


@pytest.fixture
def sample_signals():
    return [
        Signal(
            rule_id="R205", phase=2, sheet="Modele", refs=["D35", "I35"],
            label="Produits (report)", observed="=SUM($D6:I9)", expected="=SUM(I6:I9)",
            evidence={
                "chiffrage": "actif",
                "ecart_max_absolu": 5400.0,
                "colonnes_touchees": ["E35", "F35"],
                "cache_confirme_la_formule": True,
                "detail_colonnes": [
                    {
                        "colonne": "E", "ref": "E35", "cache": 2040.0,
                        "recalcul_formule_ecrite": 2040.0,
                        "recalcul_formule_attendue": 1040.0,
                        "ecart": -1000.0, "statut": "ecart",
                    }
                ],
            },
            consumers=["Modele!41"], confidence=1.0,
            note="Bornes ancrees de facon asymetrique.",
        ),
        Signal(
            rule_id="R215", phase=2, sheet="Modele", refs=["49:49"],
            label="Impasse de tresorerie", observed="=D18-D16",
            evidence={"canal_debranche": True}, consumers=[], confidence=0.85,
            note="Ligne calculee et consommee par personne.",
        ),
    ]


@pytest.fixture
def sample_log():
    log = RunLog()
    log.start("modele.xlsm", "abc123")
    log.add(RuleRun("R205", "Ancrage asymetrique", 2, "executed", signals=1))
    log.add(RuleRun("R301", "Bouclages", 3, "skipped", reason="aucun mapping fourni"))
    log.reserves = [
        Reserve("CALCPR_0", 'calcOnSave="0" : valeurs en cache non convergees', "warning")
    ]
    log.finish()
    return log


class TestFormulaTextGuard:
    """Une formule citee ne doit jamais redevenir une formule."""

    def test_write_text_forces_string_type(self):
        wb = openpyxl.Workbook()
        cell = wb.active.cell(row=1, column=1)
        write_text(cell, "=SUM(A1:A9)")
        assert cell.data_type == "s"
        assert cell.value == "=SUM(A1:A9)"

    @pytest.mark.parametrize("value", ["=A1", "+A1", "-A1", "@SUM(A1)"])
    def test_all_formula_leaders_are_neutralised(self, value):
        wb = openpyxl.Workbook()
        cell = wb.active.cell(row=1, column=1)
        write_text(cell, value)
        assert cell.data_type == "s"

    def test_negative_numbers_stay_readable(self):
        wb = openpyxl.Workbook()
        cell = wb.active.cell(row=1, column=1)
        write_text(cell, "-12.5")
        assert cell.value == "-12.5"

    def test_report_contains_no_live_formula(self, sample_signals, sample_log, tmp_path):
        out = write_report(sample_signals, tmp_path / "constats.xlsx", sample_log)
        with zipfile.ZipFile(out) as zf:
            tags = sum(
                len(re.findall(rb"<f[ >]", zf.read(name)))
                for name in zf.namelist()
                if name.startswith("xl/worksheets/")
            )
        assert tags == 0, "le classeur de constats ne doit contenir aucune formule"

    def test_cited_formulas_are_present_as_text(self, sample_signals, sample_log, tmp_path):
        out = write_report(sample_signals, tmp_path / "constats.xlsx", sample_log)
        wb = openpyxl.load_workbook(out)
        cells = [c.value for row in wb["Signaux"].iter_rows() for c in row]
        assert "=SUM($D6:I9)" in cells
        assert "=SUM(I6:I9)" in cells


class TestXlsxReport:
    def test_sheets_are_present(self, sample_signals, sample_log, tmp_path):
        out = write_report(sample_signals, tmp_path / "r.xlsx", sample_log)
        wb = openpyxl.load_workbook(out)
        assert set(wb.sheetnames) == {"Signaux", "Reserves", "Journal", "Evidence"}

    def test_signals_are_sorted_by_confidence(self, sample_signals, sample_log, tmp_path):
        out = write_report(sample_signals, tmp_path / "r.xlsx", sample_log)
        ws = openpyxl.load_workbook(out)["Signaux"]
        assert ws.cell(row=2, column=1).value == "R205"
        assert ws.cell(row=2, column=3).value == 1.0

    def test_reserves_are_carried_over(self, sample_signals, sample_log, tmp_path):
        out = write_report(sample_signals, tmp_path / "r.xlsx", sample_log)
        ws = openpyxl.load_workbook(out)["Reserves"]
        assert ws.cell(row=2, column=1).value == "CALCPR_0"

    def test_skipped_rules_appear_in_the_log(self, sample_signals, sample_log, tmp_path):
        out = write_report(sample_signals, tmp_path / "r.xlsx", sample_log)
        ws = openpyxl.load_workbook(out)["Journal"]
        rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2, max_row=3)]
        assert any(r[3] == "skipped" and "mapping" in (r[6] or "") for r in rows)

    def test_evidence_sheet_carries_the_column_detail(self, sample_signals, sample_log, tmp_path):
        out = write_report(sample_signals, tmp_path / "r.xlsx", sample_log)
        ws = openpyxl.load_workbook(out)["Evidence"]
        assert ws.cell(row=2, column=3).value == "E35"
        assert ws.cell(row=2, column=9).value == pytest.approx(-1000.0)


class TestMarkdown:
    def test_contains_the_boundary_statement(self, sample_signals, sample_log):
        text = render_markdown(sample_signals, sample_log)
        assert "ne contient aucun constat d'audit" in text

    def test_reserves_precede_the_signals(self, sample_signals, sample_log):
        text = render_markdown(sample_signals, sample_log)
        assert text.index("Reserves methodologiques") < text.index("## Signaux")

    def test_formulas_are_fenced(self, sample_signals, sample_log):
        text = render_markdown(sample_signals, sample_log)
        assert "`=SUM($D6:I9)`" in text

    def test_quantification_is_reported(self, sample_signals, sample_log):
        text = render_markdown(sample_signals, sample_log)
        assert "**actif**" in text
        assert "5400" in text.replace(" ", "")

    def test_skipped_rules_are_declared(self, sample_signals, sample_log):
        text = render_markdown(sample_signals, sample_log)
        assert "Controles non executes" in text
        assert "aucun mapping fourni" in text

    def test_orphan_states_absence_of_consumers(self, sample_signals, sample_log):
        text = render_markdown(sample_signals, sample_log)
        assert "Consommateurs : **aucun**" in text


class TestGraphviz:
    def test_mermaid_ids_are_safe(self, healthy_ctx):
        graph = healthy_ctx.graph
        text = sheet_map_mermaid(graph)
        assert "graph LR" in text
        # Un nom d'onglet avec espaces ne doit pas casser la syntaxe Mermaid.
        assert "Etats_fi_periode[" in text
        assert "Etats fi periode[" not in text

    def test_dot_output(self, healthy_ctx):
        text = sheet_map_dot(healthy_ctx.graph)
        assert text.startswith("digraph")
        assert '"Modele"' in text

    def test_trace_tree(self, healthy_ctx):
        node = healthy_ctx.graph.trace("Modele", 18, 4, depth=1, direction="up")
        text = trace_mermaid(node, "up")
        assert "graph BT" in text
        assert "Modele_D18" in text


class TestJsonContract:
    def test_document_is_versioned_and_sorted(self, sample_signals, sample_log):
        doc = signals_to_document(sample_signals, sample_log)
        assert doc["schema_version"] == "1.0"
        assert doc["signal_count"] == 2
        confidences = [s["confidence"] for s in doc["signals"]]
        assert confidences == sorted(confidences, reverse=True)

    def test_document_is_json_serialisable(self, sample_signals, sample_log):
        text = json.dumps(signals_to_document(sample_signals, sample_log))
        assert "R205" in text

    def test_signals_round_trip_back_into_objects(self, sample_signals, sample_log):
        doc = signals_to_document(sample_signals, sample_log)
        restored = [Signal(**item) for item in doc["signals"]]
        assert restored[0].rule_id == "R205"
        assert restored[0].evidence["chiffrage"] == "actif"

    def test_run_log_lists_skipped_rules(self, sample_log):
        doc = signals_to_document([], sample_log)
        skipped = [r for r in doc["run"]["rules"] if r["status"] == "skipped"]
        assert skipped and skipped[0]["reason"] == "aucun mapping fourni"
