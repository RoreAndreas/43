"""Classeur de constats.

Piege traite ici : une cellule qui recoit une chaine commencant par `=` est
interpretee par Excel comme une formule. Un rapport qui cite des formules — ce
que fait chaque signal, deux fois — produirait alors un classeur truffe de
formules actives, de `#REF!` et de references vers des cellules du rapport
lui-meme. Toute valeur commencant par `=` passe donc par `write_text`, qui force
`cell.data_type = "s"`.

Le fichier source n'est jamais touche : ce module ecrit un classeur neuf.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import RunLog, Signal

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONO = Font(name="Consolas", size=9)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")

COLUMNS = [
    ("Regle", 8),
    ("Phase", 6),
    ("Confiance", 10),
    ("Onglet", 22),
    ("References", 18),
    ("Libelle", 34),
    ("Constat", 60),
    ("Formule relevee", 44),
    ("Formule attendue", 44),
    ("Chiffrage", 12),
    ("Ecart max", 14),
    ("Colonnes touchees", 24),
    ("Consommateurs", 30),
]


def write_text(cell, value: Any) -> None:
    """Ecrit une valeur en forcant le type texte quand elle ressemble a une formule.

    Sans ce garde-fou, `=SUM(D6:D9)` cite dans un rapport devient une formule
    vivante a l'ouverture du classeur.
    """
    if value is None:
        cell.value = None
        return
    text = value if isinstance(value, str) else str(value)
    stripped = text.lstrip()
    if stripped.startswith(("=", "+", "-", "@")) and not _is_plain_number(stripped):
        cell.value = text
        cell.data_type = "s"
    else:
        cell.value = text


def _is_plain_number(text: str) -> bool:
    try:
        float(text)
        return True
    except ValueError:
        return False


def _confidence_fill(confidence: float) -> PatternFill:
    """Bandeau de tri visuel. Ne remplace pas la valeur, qui reste lisible."""
    if confidence >= 0.85:
        return PatternFill("solid", fgColor="F8CBAD")
    if confidence >= 0.6:
        return PatternFill("solid", fgColor="FFE699")
    return PatternFill("solid", fgColor="E2EFDA")


def write_report(
    signals: list[Signal],
    path: str | Path,
    run_log: RunLog | None = None,
) -> Path:
    """Ecrit le classeur de constats : signaux, reserves, journal."""
    path = Path(path)
    wb = Workbook()

    ws = wb.active
    ws.title = "Signaux"
    _write_signals(ws, signals)
    _write_reserves(wb.create_sheet("Reserves"), run_log)
    _write_log(wb.create_sheet("Journal"), run_log)
    _write_evidence(wb.create_sheet("Evidence"), signals)

    wb.save(path)
    wb.close()
    return path


def _header(ws, columns: list[tuple[str, int]]) -> None:
    for i, (title, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _write_signals(ws, signals: list[Signal]) -> None:
    _header(ws, COLUMNS)
    ordered = sorted(signals, key=lambda s: (-s.confidence, s.rule_id, s.sheet))
    for r, sig in enumerate(ordered, start=2):
        evidence = sig.evidence or {}
        row_values = [
            sig.rule_id,
            sig.phase,
            round(sig.confidence, 2),
            sig.sheet,
            ", ".join(sig.refs),
            sig.label,
            sig.note,
            sig.observed,
            sig.expected,
            evidence.get("chiffrage"),
            evidence.get("ecart_max_absolu"),
            ", ".join(evidence.get("colonnes_touchees", [])[:12]),
            ", ".join(sig.consumers[:8]),
        ]
        for c, value in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c)
            if c in (2, 3, 11) and isinstance(value, (int, float)):
                cell.value = value
                cell.alignment = TOP
            else:
                write_text(cell, value)
                cell.alignment = WRAP if c in (7, 8, 9, 12, 13) else TOP
            if c in (8, 9):
                cell.font = MONO
        ws.cell(row=r, column=3).fill = _confidence_fill(sig.confidence)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(1, len(ordered) + 1)}"


def _write_reserves(ws, run_log: RunLog | None) -> None:
    _header(ws, [("Code", 28), ("Gravite", 12), ("Reserve", 120)])
    reserves = run_log.reserves if run_log else []
    if not reserves:
        write_text(ws.cell(row=2, column=1), "(aucune)")
        return
    for r, reserve in enumerate(reserves, start=2):
        write_text(ws.cell(row=r, column=1), reserve.code)
        write_text(ws.cell(row=r, column=2), reserve.severity)
        cell = ws.cell(row=r, column=3)
        write_text(cell, reserve.message)
        cell.alignment = WRAP


def _write_log(ws, run_log: RunLog | None) -> None:
    _header(ws, [("Regle", 10), ("Libelle", 34), ("Phase", 7), ("Statut", 12),
                 ("Signaux", 10), ("Duree (s)", 11), ("Motif", 70)])
    if run_log is None:
        return
    for r, run in enumerate(run_log.rules, start=2):
        write_text(ws.cell(row=r, column=1), run.rule_id)
        write_text(ws.cell(row=r, column=2), run.label)
        ws.cell(row=r, column=3, value=run.phase)
        write_text(ws.cell(row=r, column=4), run.status)
        ws.cell(row=r, column=5, value=run.signals)
        ws.cell(row=r, column=6, value=run.duration_s)
        write_text(ws.cell(row=r, column=7), run.reason)

    base = len(run_log.rules) + 3
    for i, (key, value) in enumerate(
        (
            ("Classeur", run_log.workbook),
            ("SHA-256", run_log.sha256),
            ("Debut", run_log.started_at),
            ("Fin", run_log.finished_at),
            ("Version outil", run_log.tool_version),
            ("Version de schema", run_log.schema_version),
        )
    ):
        ws.cell(row=base + i, column=1, value=key).font = Font(bold=True)
        write_text(ws.cell(row=base + i, column=2), value)


def _write_evidence(ws, signals: list[Signal]) -> None:
    """Detail chiffre, une ligne par colonne analysee, pour les regles de plages."""
    _header(ws, [("Regle", 8), ("Onglet", 20), ("Reference", 14), ("Libelle", 30),
                 ("Colonne", 10), ("Cache", 16), ("Recalcul ecrit", 16),
                 ("Recalcul attendu", 18), ("Ecart", 16), ("Statut", 12)])
    r = 2
    for sig in sorted(signals, key=lambda s: -s.confidence):
        detail = (sig.evidence or {}).get("detail_colonnes") or []
        for entry in detail:
            if entry.get("ecart") is None:
                continue
            write_text(ws.cell(row=r, column=1), sig.rule_id)
            write_text(ws.cell(row=r, column=2), sig.sheet)
            write_text(ws.cell(row=r, column=3), entry.get("ref"))
            write_text(ws.cell(row=r, column=4), sig.label)
            write_text(ws.cell(row=r, column=5), entry.get("colonne"))
            for col, key in (
                (6, "cache"),
                (7, "recalcul_formule_ecrite"),
                (8, "recalcul_formule_attendue"),
                (9, "ecart"),
            ):
                value = entry.get(key)
                if isinstance(value, (int, float)):
                    ws.cell(row=r, column=col, value=value)
                else:
                    write_text(ws.cell(row=r, column=col), value)
            write_text(ws.cell(row=r, column=10), entry.get("statut"))
            r += 1
    if r == 2:
        write_text(ws.cell(row=2, column=1), "(aucun chiffrage detaille)")
