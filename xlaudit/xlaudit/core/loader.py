"""Chargement du classeur, iteration sure, cache disque.

Trois contraintes structurantes, toutes issues de degats constates :

1. Double chargement. openpyxl ne rend jamais formules et valeurs en meme temps :
   il faut deux `load_workbook`, l'un `data_only=False`, l'autre `data_only=True`.

2. `read_only=True` obligatoire. Au-dela de quelques centaines de milliers de
   formules, le mode normal sature la memoire.

3. Piege du mode lecture seule : les cellules vides sont des `EmptyCell`, sans
   attribut `.row` ni `.column`. Tout code qui fait `cell.row` dans une boucle
   `iter_rows` plante des qu'une ligne comporte un trou.

   Verifie empiriquement sur openpyxl 3.1 : `iter_rows()` sans bornes commence a
   la ligne 1 et a la colonne 1, *pas* a `ws.min_row` / `ws.min_column`. Un
   `enumerate(row, start=ws.min_column)` serait donc faux. La parade retenue :
   toujours passer des bornes explicites a `iter_rows`, puis deduire la ligne d'un
   compteur et la colonne d'un `enumerate(row, start=min_col)`. Les lignes
   entierement vides sont bien restituees (padding), donc le compteur ne derive
   jamais ; l'iteration s'arrete en revanche a la derniere ligne reelle, ce qui
   raccourcit la boucle sans la desynchroniser.

`iter_cells` est le seul endroit du projet autorise a appeler `iter_rows`. Un test
(`tests/test_no_direct_iter_rows.py`) fait respecter cette regle.

Aucune fonction de ce module n'ouvre le classeur en ecriture. Le fichier source
n'est jamais modifie, sous aucune option.
"""

from __future__ import annotations

import gzip
import hashlib
import json
import os
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Iterator
from xml.etree import ElementTree as ET

import openpyxl

from ..models import SCHEMA_VERSION, _jsonable
from . import refs as R

CACHE_DIRNAME = ".xlaudit_cache"
_NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

ProgressCb = Callable[[str, int, int], None]


def sha256_file(path: str | os.PathLike, chunk: int = 1 << 20) -> str:
    """Empreinte du classeur : sert de cle de cache et figure au journal."""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        while True:
            block = fh.read(chunk)
            if not block:
                break
            h.update(block)
    return h.hexdigest()


def normalize_cell_value(value: Any) -> Any:
    """Normalise la valeur brute d'une cellule.

    Une formule matricielle remonte comme un objet `ArrayFormula`, pas comme une
    chaine : sans ce `getattr(v, "text", v)` les formules matricielles
    disparaissent silencieusement des scans. Meme traitement pour les tables de
    donnees (`DataTableFormula`).
    """
    if value is None:
        return None
    text = getattr(value, "text", None)
    if text is not None:
        return text
    if type(value).__name__ == "DataTableFormula":
        return "=TABLE()"
    return value


def iter_cells(
    ws,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
    skip_empty: bool = True,
) -> Iterator[tuple[int, int, Any]]:
    """Iteration sure sur les cellules d'un onglet en lecture seule.

    Rend des triplets (ligne, colonne, valeur normalisee) sans jamais toucher
    `cell.row` ni `cell.column`, qui n'existent pas sur les cellules vides.

    Seul point du projet ou `iter_rows` est appele.
    """
    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = ws.max_column
    if not max_row or not max_col:
        return
    if max_row < min_row or max_col < min_col:
        return

    row_idx = min_row - 1
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        row_idx += 1
        for col_idx, cell in enumerate(row, start=min_col):
            value = normalize_cell_value(getattr(cell, "value", None))
            if value is None and skip_empty:
                continue
            if skip_empty and isinstance(value, str) and not value:
                continue
            yield row_idx, col_idx, value


@dataclass
class SheetSnapshot:
    """Etat collecte d'un onglet : formules, valeurs en cache, libelles."""

    name: str
    index: int = 0
    state: str = "visible"
    min_row: int = 1
    max_row: int = 1
    min_col: int = 1
    max_col: int = 1
    formulas: dict[tuple[int, int], str] = field(default_factory=dict)
    values: dict[tuple[int, int], Any] = field(default_factory=dict)
    label_max_col: int = 12

    # -- acces -------------------------------------------------------------
    def formula(self, row: int, col: int) -> str | None:
        return self.formulas.get((row, col))

    def value(self, row: int, col: int) -> Any:
        return self.values.get((row, col))

    def formula_cols(self, row: int) -> list[int]:
        return sorted(c for (r, c) in self.formulas if r == row)

    def has_content(self, row: int, col: int) -> bool:
        return (row, col) in self.formulas or (row, col) in self.values

    def row_label(self, row: int) -> str:
        """Libelle de la ligne : texte le plus a droite dans la zone de libelles.

        Dans un modele financier les libelles sont indentes sur plusieurs colonnes ;
        le plus a droite est le plus specifique (`   dont interets` plutot que
        `Charges financieres`). On ignore les cellules porteuses de formule, qui
        ne sont pas des libelles.
        """
        best = ""
        best_col = -1
        for col in range(1, self.label_max_col + 1):
            if (row, col) in self.formulas:
                continue
            v = self.values.get((row, col))
            if isinstance(v, str) and v.strip():
                if col > best_col:
                    best, best_col = v.strip(), col
        return best

    def to_json(self) -> dict:
        return {
            "name": self.name,
            "index": self.index,
            "state": self.state,
            "min_row": self.min_row,
            "max_row": self.max_row,
            "min_col": self.min_col,
            "max_col": self.max_col,
            "label_max_col": self.label_max_col,
            "formulas": [[r, c, f] for (r, c), f in self.formulas.items()],
            "values": [[r, c, _jsonable(v)] for (r, c), v in self.values.items()],
        }

    @classmethod
    def from_json(cls, d: dict) -> "SheetSnapshot":
        s = cls(
            name=d["name"],
            index=d.get("index", 0),
            state=d.get("state", "visible"),
            min_row=d.get("min_row", 1),
            max_row=d.get("max_row", 1),
            min_col=d.get("min_col", 1),
            max_col=d.get("max_col", 1),
            label_max_col=d.get("label_max_col", 12),
        )
        s.formulas = {(r, c): f for r, c, f in d.get("formulas", [])}
        s.values = {(r, c): v for r, c, v in d.get("values", [])}
        return s


@dataclass
class DefinedName:
    name: str
    scope: str | None  # None = portee globale, sinon nom de l'onglet
    refers_to: str = ""
    hidden: bool = False

    def to_json(self) -> dict:
        return {
            "name": self.name,
            "scope": self.scope,
            "refers_to": self.refers_to,
            "hidden": self.hidden,
        }


@dataclass
class Snapshot:
    """Resultat du passage de collecte unique.

    Toutes les analyses repartent de cet objet : le classeur n'est jamais relu.
    """

    path: str = ""
    sha256: str = ""
    sheets: dict[str, SheetSnapshot] = field(default_factory=dict)
    defined_names: list[DefinedName] = field(default_factory=list)
    calc: dict = field(default_factory=dict)
    vba: dict = field(default_factory=dict)
    stats: dict = field(default_factory=dict)

    # -- acces -------------------------------------------------------------
    @property
    def sheet_names(self) -> list[str]:
        return [s.name for s in sorted(self.sheets.values(), key=lambda x: x.index)]

    def sheet(self, name: str) -> SheetSnapshot | None:
        if name in self.sheets:
            return self.sheets[name]
        # Excel est insensible a la casse sur les noms d'onglet.
        low = name.lower()
        for s in self.sheets.values():
            if s.name.lower() == low:
                return s
        return None

    def formula(self, sheet: str, row: int, col: int) -> str | None:
        s = self.sheet(sheet)
        return s.formula(row, col) if s else None

    def value(self, sheet: str, row: int, col: int) -> Any:
        s = self.sheet(sheet)
        return s.value(row, col) if s else None

    def row_label(self, sheet: str, row: int) -> str:
        s = self.sheet(sheet)
        return s.row_label(row) if s else ""

    @property
    def formula_count(self) -> int:
        return sum(len(s.formulas) for s in self.sheets.values())

    # -- serialisation -----------------------------------------------------
    def to_json(self) -> dict:
        return {
            "schema_version": SCHEMA_VERSION,
            "kind": "xlaudit-snapshot",
            "path": self.path,
            "sha256": self.sha256,
            "calc": self.calc,
            "vba": self.vba,
            "stats": self.stats,
            "defined_names": [d.to_json() for d in self.defined_names],
            "sheets": [s.to_json() for s in self.sheets.values()],
        }

    @classmethod
    def from_json(cls, d: dict) -> "Snapshot":
        snap = cls(
            path=d.get("path", ""),
            sha256=d.get("sha256", ""),
            calc=d.get("calc", {}),
            vba=d.get("vba", {}),
            stats=d.get("stats", {}),
        )
        snap.defined_names = [
            DefinedName(
                name=x["name"], scope=x.get("scope"),
                refers_to=x.get("refers_to", ""), hidden=x.get("hidden", False),
            )
            for x in d.get("defined_names", [])
        ]
        for sd in d.get("sheets", []):
            sheet = SheetSnapshot.from_json(sd)
            snap.sheets[sheet.name] = sheet
        return snap

    def save(self, path: str | os.PathLike) -> None:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with gzip.open(path, "wt", encoding="utf-8") as fh:
            json.dump(self.to_json(), fh, ensure_ascii=False)

    @classmethod
    def load(cls, path: str | os.PathLike) -> "Snapshot":
        with gzip.open(path, "rt", encoding="utf-8") as fh:
            return cls.from_json(json.load(fh))


# ---------------------------------------------------------------------------
# Lecture des metadonnees XML : calcPr et noms definis
# ---------------------------------------------------------------------------

def read_calc_properties(path: str | os.PathLike) -> dict:
    """Lit `calcPr` dans xl/workbook.xml.

    `calcOnSave="0"` signifie que les valeurs en cache sont celles du dernier
    calcul de l'utilisateur, pas l'etat converge du modele : tout chiffrage
    adosse au cache herite d'une reserve. `iterate="1"` signale que le classeur
    resout des circularites par iteration.
    """
    out: dict = {}
    try:
        with zipfile.ZipFile(path) as zf:
            with zf.open("xl/workbook.xml") as fh:
                root = ET.parse(fh).getroot()
    except (KeyError, zipfile.BadZipFile, ET.ParseError, FileNotFoundError):
        return out
    node = root.find(f"{_NS_MAIN}calcPr")
    if node is None:
        return out
    out = dict(node.attrib)
    return out


def read_defined_names_xml(path: str | os.PathLike) -> list[DefinedName]:
    """Lit tous les noms definis, portee globale *et* portee feuille.

    `wb.defined_names` ne contient que la portee globale : les noms a portee
    feuille portent un attribut `localSheetId` et resteraient invisibles, soit
    souvent la moitie du gestionnaire de noms.
    """
    names: list[DefinedName] = []
    try:
        with zipfile.ZipFile(path) as zf:
            with zf.open("xl/workbook.xml") as fh:
                root = ET.parse(fh).getroot()
    except (KeyError, zipfile.BadZipFile, ET.ParseError, FileNotFoundError):
        return names

    sheet_order = [
        s.get("name", "") for s in root.findall(f"{_NS_MAIN}sheets/{_NS_MAIN}sheet")
    ]
    container = root.find(f"{_NS_MAIN}definedNames")
    if container is None:
        return names
    for node in container.findall(f"{_NS_MAIN}definedName"):
        local = node.get("localSheetId")
        scope = None
        if local is not None:
            try:
                scope = sheet_order[int(local)]
            except (ValueError, IndexError):
                scope = None
        names.append(
            DefinedName(
                name=node.get("name", ""),
                scope=scope,
                refers_to=(node.text or "").strip(),
                hidden=node.get("hidden") == "1",
            )
        )
    return names


def _sheet_states(path: str | os.PathLike) -> dict[str, str]:
    """Etat visible / hidden / veryHidden de chaque onglet."""
    out: dict[str, str] = {}
    try:
        with zipfile.ZipFile(path) as zf:
            with zf.open("xl/workbook.xml") as fh:
                root = ET.parse(fh).getroot()
    except (KeyError, zipfile.BadZipFile, ET.ParseError, FileNotFoundError):
        return out
    for s in root.findall(f"{_NS_MAIN}sheets/{_NS_MAIN}sheet"):
        out[s.get("name", "")] = s.get("state", "visible")
    return out


# ---------------------------------------------------------------------------
# Passage de collecte
# ---------------------------------------------------------------------------

def _guess_label_max_col(formula_cols: dict[int, int], max_col: int) -> int:
    """Derniere colonne de la zone de libelles.

    Les colonnes de projection concentrent les formules ; la zone de libelles est
    ce qui les precede. On prend la premiere colonne ou la densite de formules
    decolle, moins un, borne entre 1 et 26.
    """
    if not formula_cols:
        return min(12, max(1, max_col))
    peak = max(formula_cols.values())
    if peak <= 0:
        return min(12, max(1, max_col))
    for col in range(1, min(max_col, 40) + 1):
        if formula_cols.get(col, 0) >= 0.3 * peak:
            return max(1, min(col - 1, 26))
    return min(12, max(1, max_col))


def collect(
    path: str | os.PathLike,
    progress: ProgressCb | None = None,
    max_sheets: int | None = None,
) -> Snapshot:
    """Passage de collecte unique : un seul parcours alimente tous les controles.

    Ouvre le classeur deux fois en lecture seule (formules puis valeurs) et ne
    l'ecrit jamais.
    """
    path = Path(path)
    snap = Snapshot(path=str(path), sha256=sha256_file(path))
    snap.calc = read_calc_properties(path)
    snap.defined_names = read_defined_names_xml(path)
    states = _sheet_states(path)

    wb_f = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        names = list(wb_f.sheetnames)
        if max_sheets:
            names = names[:max_sheets]
        total = len(names)
        for i, name in enumerate(names):
            if progress:
                progress(f"formules · {name}", i, total)
            ws = wb_f[name]
            sheet = SheetSnapshot(
                name=name,
                index=i,
                state=states.get(name, "visible"),
                min_row=ws.min_row or 1,
                max_row=ws.max_row or 1,
                min_col=ws.min_column or 1,
                max_col=ws.max_column or 1,
            )
            per_col: dict[int, int] = {}
            for r, c, v in iter_cells(ws, min_row=1, min_col=1):
                if isinstance(v, str) and v.startswith("="):
                    sheet.formulas[(r, c)] = v
                    per_col[c] = per_col.get(c, 0) + 1
                else:
                    # Constantes : conservees ici, elles seront ecrasees par les
                    # valeurs en cache du second chargement si elles y figurent.
                    sheet.values[(r, c)] = v
            sheet.label_max_col = _guess_label_max_col(per_col, sheet.max_col)
            snap.sheets[name] = sheet
        if progress:
            progress("formules", total, total)
    finally:
        wb_f.close()

    wb_v = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = [n for n in wb_v.sheetnames if n in snap.sheets]
        total = len(names)
        cached_any = False
        for i, name in enumerate(names):
            if progress:
                progress(f"valeurs · {name}", i, total)
            ws = wb_v[name]
            sheet = snap.sheets[name]
            for r, c, v in iter_cells(ws, min_row=1, min_col=1):
                sheet.values[(r, c)] = v
                if (r, c) in sheet.formulas:
                    cached_any = True
        if progress:
            progress("valeurs", total, total)
    finally:
        wb_v.close()

    snap.stats = {
        "sheets": len(snap.sheets),
        "formulas": snap.formula_count,
        "values": sum(len(s.values) for s in snap.sheets.values()),
        "defined_names": len(snap.defined_names),
        "has_cached_values": cached_any,
    }
    return snap


def cache_path_for(path: str | os.PathLike, sha: str, cache_dir: str | os.PathLike | None = None) -> Path:
    base = Path(cache_dir) if cache_dir else Path(path).resolve().parent / CACHE_DIRNAME
    return base / f"{Path(path).stem}.{sha[:16]}.snapshot.json.gz"


def get_snapshot(
    path: str | os.PathLike,
    cache_dir: str | os.PathLike | None = None,
    force: bool = False,
    use_cache: bool = True,
    progress: ProgressCb | None = None,
) -> tuple[Snapshot, bool]:
    """Rend le snapshot, depuis le cache si l'empreinte du classeur correspond.

    Le cache est invalide par le SHA-256 du fichier : il figure dans le nom du
    fichier de cache, donc une modification du classeur produit une cle
    differente et ne peut pas etre servie par erreur.

    Rend (snapshot, depuis_cache).
    """
    sha = sha256_file(path)
    cpath = cache_path_for(path, sha, cache_dir)
    if use_cache and not force and cpath.exists():
        try:
            snap = Snapshot.load(cpath)
            if snap.sha256 == sha:
                return snap, True
        except (OSError, ValueError, KeyError, EOFError):
            pass  # cache illisible : on recollecte
    snap = collect(path, progress=progress)
    if use_cache:
        try:
            snap.save(cpath)
        except OSError:
            pass  # un cache non ecrit ne doit jamais faire echouer une analyse
    return snap, False


def load_vba(path: str | os.PathLike) -> dict:
    """Extraction VBA deportee dans core.vba (import tardif : oletools est lourd)."""
    from .vba import extract_vba

    return extract_vba(path)
