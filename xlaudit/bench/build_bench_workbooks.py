"""Construit les classeurs du benchmark des moteurs d'evaluation.

Deux classeurs :

- `couverture.xlsx` : petit, mais contenant chacune des fonctions et structures
  citees par la specification (XLOOKUP, LET, FILTER, UNIQUE, SUMIFS, formule
  matricielle, circularite simple). Chaque cellule a une **valeur attendue
  connue**, calculee ici en Python, ce qui permet de distinguer les trois
  comportements possibles d'un moteur : bonne valeur, echec explicite, ou
  *valeur fausse rendue en silence* — le seul qui soit eliminatoire.

- `charge.xlsx` : 200 000 formules sur 20 onglets, pour mesurer le temps de
  chargement et l'empreinte memoire.

Les valeurs en cache sont injectees dans le XML comme le ferait Excel, sans quoi
il serait impossible de comparer la sortie d'un moteur a la verite.
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from tests.fixtures import inject_cached_values

SHEET = "Couverture"

# (ref, formule, valeur attendue, etiquette de la fonction testee)
CASES: list[tuple[str, str, object, str]] = [
    ("D2", "=SUM(B2:B6)", 150.0, "SUM"),
    # cles x,y,x,y,x et valeurs 10..50 -> les lignes « x » valent 10+30+50 = 90.
    ("D3", '=SUMIFS(B2:B6,A2:A6,"x")', 90.0, "SUMIFS"),
    ("D4", "=INDEX(B2:B6,2)", 20.0, "INDEX"),
    ("D5", "=MATCH(30,B2:B6,0)", 3.0, "MATCH"),
    ("D6", "=VLOOKUP(30,B2:C6,2,FALSE)", 300.0, "VLOOKUP"),
    ("D7", "=IFERROR(1/0,-1)", -1.0, "IFERROR"),
    ("D8", "=ROUND(SUM(B2:B6)/7,2)", 21.43, "ROUND"),
    # Fonctions modernes : le coeur du test.
    ("D9", "=_xlfn.XLOOKUP(30,B2:B6,C2:C6)", 300.0, "XLOOKUP"),
    ("D10", "=_xlfn.LET(x,SUM(B2:B6),x*2)", 300.0, "LET"),
    # valeurs strictement superieures a 20 : 30 + 40 + 50 = 120.
    ("D11", "=SUM(_xlfn._xlws.FILTER(B2:B6,B2:B6>20))", 120.0, "FILTER"),
    ("D12", "=SUM(_xlfn.UNIQUE(B2:B6))", 150.0, "UNIQUE"),
    ("D13", "=_xlfn.SEQUENCE(1)", 1.0, "SEQUENCE"),
    ("D14", "=_xlfn.TEXTJOIN(\",\",TRUE,A2:A3)", "x,y", "TEXTJOIN"),
]

# Circularite simple : E2 et E3 se lisent mutuellement.
# Point fixe : E2 = E3/2 + 10 et E3 = E2 + 5  ->  E2 = 25, E3 = 30.
CIRCULAR = [("E2", "=E3/2+10", 25.0), ("E3", "=E2+5", 30.0)]


def build_coverage(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET

    ws["A1"], ws["B1"], ws["C1"] = "cle", "valeur", "assoc"
    data = [("x", 10.0, 100.0), ("y", 20.0, 200.0), ("x", 30.0, 300.0),
            ("y", 40.0, 400.0), ("x", 50.0, 500.0)]
    for i, (k, v, a) in enumerate(data, start=2):
        ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"] = k, v, a

    cached: dict[tuple[int, int], object] = {}
    for ref, formula, expected, _name in CASES:
        ws[ref] = formula
        cell = ws[ref]
        cached[(cell.row, cell.column)] = expected
    for ref, formula, expected in CIRCULAR:
        ws[ref] = formula
        cell = ws[ref]
        cached[(cell.row, cell.column)] = expected

    # Formule matricielle : la valeur remonte comme ArrayFormula, pas comme texte.
    ws["G2"] = ArrayFormula("G2", "=SUM(B2:B6*C2:C6)")
    cached[(2, 7)] = 55000.0

    wb.save(path)
    wb.close()
    inject_cached_values(path, {SHEET: cached})
    return path


def build_load(path: Path, n_sheets: int = 20, n_rows: int = 500, n_cols: int = 20) -> Path:
    """200 000 formules reparties sur 20 onglets."""
    wb = openpyxl.Workbook(write_only=True)
    for s in range(n_sheets):
        ws = wb.create_sheet(f"Onglet{s}")
        for r in range(1, n_rows + 1):
            row: list = [f"Poste {r}"]
            for c in range(n_cols):
                letter = get_column_letter(c + 2)
                if r <= 5:
                    row.append(float(r * 10 + c))
                elif r % 17 == 0:
                    row.append(f"=SUM({letter}{r - 10}:{letter}{r - 1})")
                else:
                    row.append(f"={letter}{r - 1}*1.02+{letter}5")
            ws.append(row)
    wb.save(path)
    wb.close()
    return path


if __name__ == "__main__":
    out = Path(sys.argv[1] if len(sys.argv) > 1 else ".")
    out.mkdir(parents=True, exist_ok=True)
    c = build_coverage(out / "couverture.xlsx")
    print(f"couverture : {c} ({len(CASES)} fonctions + circularite + matricielle)")
    l = build_load(out / "charge.xlsx")
    import openpyxl as op

    wb = op.load_workbook(l, read_only=True)
    print(f"charge     : {l} ({len(wb.sheetnames)} onglets)")
    wb.close()
