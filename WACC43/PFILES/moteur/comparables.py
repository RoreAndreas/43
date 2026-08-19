"""
Univers de comparables, lu depuis l'export S&P Global déposé dans WACC43.

L'export tient en trois onglets qui partagent la même liste de sociétés, dans le
même ordre, appariés par `Entity ID` :

    Sheet1  produits, EBITDA, marge, résultat net (FY2022 à FY2025),
            puis bêta 1 an (colonne S) et bêta 3 ans (colonne T)
    Sheet2  géographie, pays, industrie, industrie primaire, description
    Sheet3  dette totale, capitalisation boursière

Seules les sociétés **complètes** alimentent la plateforme : pays, industrie,
les deux bêtas, dette, capitalisation, chiffre d'affaires sur quatre exercices
et résultat net sur trois. Les autres sont conservées dans l'univers mais
marquées `visible: False`, et n'entrent ni dans les médianes ni dans les
effectifs affichés.

L'EBITDA est délibérément hors critère. S&P n'en publie pas pour les banques ni
les assureurs — la notion n'a pas de sens pour elles — et l'exiger ferait tomber
le secteur bancaire de 112 sociétés à 1, alors qu'il est le premier des places
africaines. Un critère qui détruit le secteur principal ne mesure plus la
complétude, il mesure le modèle comptable.

L'industrie retenue est la colonne E de Sheet2 (`IQ_INDUSTRY`), soit une
soixantaine de secteurs. La colonne F (`IQ_PRIMARY_INDUSTRY`) est plus fine mais
trop éclatée pour servir de maille de comparables.

Le fichier porte les deux bêtas cotés, à un an et à trois ans. C'est le trois ans
qui alimente le coût des fonds propres : sur des marchés peu liquides, celui à un
an bouge trop pour servir de base à un coût du capital. Dette et capitalisation
donnent par ailleurs un gearing sectoriel observé — celui qu'il fallait jusqu'ici
saisir à la main.

Piège d'unités : S&P exporte les comptes et la dette en **milliers** (en-tête
`BCEAO000`) mais la capitalisation en **millions** (`BCEAOM`). Rapporter l'une à
l'autre sans conversion donne un gearing mille fois trop élevé — 593 pour les
banques au lieu de 0,59. Tout est donc ramené ici en millions.

Les en-têtes occupent les lignes 3 à 6 ; les données commencent ligne 7.
"""

from __future__ import annotations

import statistics
from pathlib import Path

import openpyxl

PREMIERE_LIGNE = 7

# Nombre de sociétés retenues pour une médiane sectorielle. Voir
# `statistiques()` : au-delà, ce sont les micro-capitalisations qui décident.
ECHANTILLON_MAX = 20
EXERCICES = ["FY2025", "FY2024", "FY2023", "FY2022"]

# S&P écrit « NA » quand la donnée manque, et parfois « NM » (non significatif).
_ABSENT = {"NA", "NM", "NC", "-", ""}


def _nombre(valeur):
    if valeur is None:
        return None
    if isinstance(valeur, (int, float)):
        return float(valeur)
    texte = str(valeur).strip()
    if texte.upper() in _ABSENT:
        return None
    try:
        return float(texte.replace(",", ""))
    except ValueError:
        return None


def _nom_court(nom: str) -> str:
    """« Absa Bank Kenya PLC (NASE:ABSA) » -> « Absa Bank Kenya PLC »."""
    texte = str(nom).strip()
    if texte.endswith(")") and "(" in texte:
        return texte[: texte.rindex("(")].strip()
    return texte


def _place_et_ticker(nom: str, entity_id):
    """« Absa Group Limited (JSE:ABG) » -> ("JSE", "ABG").

    S&P préfixe le code de la valeur par celui de sa place de cotation. C'est la
    seule mention du marché dans l'export : il n'y a pas de colonne dédiée.
    """
    texte = str(nom).strip()
    if texte.endswith(")") and "(" in texte:
        dedans = texte[texte.rindex("(") + 1: -1].strip()
        if ":" in dedans:
            place, _sep, code = dedans.partition(":")
            return place.strip() or None, code.strip() or str(entity_id)
        return None, dedans or str(entity_id)
    return None, str(entity_id)


def trouver_exports(racine: Path) -> list:
    """Tous les exports `comps*.xlsx` trouvés autour du dossier fourni.

    Un fichier par grande zone — compsAF, compsEU, compsUS, et ceux à venir —
    plutôt qu'un export unique : les actualiser séparément évite de retélécharger
    vingt-cinq mille lignes américaines pour corriger une ligne africaine.

    Les exports sont déposés à la racine de WACC43, alors que le build travaille
    depuis PFILES. On regarde donc le dossier donné, son parent et son
    grand-parent : un chemin trop étroit fait passer la construction pour
    réussie tout en publiant une page sans univers de comparables, ce qui s'est
    déjà produit et n'a été vu qu'en inspectant la page en ligne.
    """
    for base in (racine, racine.parent, racine.parent.parent):
        if not base.is_dir():
            continue
        trouves = sorted(base.glob("comps*.xlsx"))
        if trouves:
            return trouves
    return []


def charger(chemin: Path, progress=None) -> dict:
    """Lit l'export et rend l'univers de comparables."""
    def dire(message):
        if progress:
            progress(message)

    dire(f"Comparables : lecture de {chemin.name}...")
    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    try:
        signaletique = {}
        for ligne in wb["Sheet2"].iter_rows(
            min_row=PREMIERE_LIGNE, min_col=1, max_col=7, values_only=True
        ):
            nom, ident, _geo, pays, industrie, primaire, description = ligne[:7]
            if not nom or not ident:
                continue
            place, code = _place_et_ticker(nom, ident)
            signaletique[str(ident)] = {
                "nom": _nom_court(nom),
                "ticker": code,
                "place": place,
                "pays": (pays or "").strip() or None,
                "industrie": (industrie or "").strip() or None,
                "industrie_fine": (primaire or "").strip() or None,
                "description": (description or "").strip() or None,
            }

        comptes = {}
        for ligne in wb["Sheet1"].iter_rows(
            min_row=PREMIERE_LIGNE, min_col=1, max_col=20, values_only=True
        ):
            ident = ligne[1]
            if not ident:
                continue
            comptes[str(ident)] = {
                "ca": {a: _nombre(ligne[2 + i]) for i, a in enumerate(EXERCICES)},
                "ebitda": {a: _nombre(ligne[6 + i]) for i, a in enumerate(EXERCICES)},
                "rn": {a: _nombre(ligne[14 + i]) for i, a in enumerate(EXERCICES[:3])},
                "beta_1an": _nombre(ligne[18]),
                "beta_3ans": _nombre(ligne[19]),
            }

        marche = {}
        for ligne in wb["Sheet3"].iter_rows(
            min_row=PREMIERE_LIGNE, min_col=1, max_col=4, values_only=True
        ):
            ident = ligne[1]
            if not ident:
                continue
            dette = _nombre(ligne[2])
            marche[str(ident)] = {
                # Dette en milliers, capitalisation en millions : on ramène tout
                # en millions pour que le rapport ait un sens.
                "dette": None if dette is None else dette / 1000.0,
                "capitalisation": _nombre(ligne[3]),
            }
    finally:
        wb.close()

    societes = {}
    for ident, base in signaletique.items():
        if not base["industrie"] or not base["pays"]:
            continue
        entree = dict(base)
        entree.update(comptes.get(ident, {}))
        entree.update(marche.get(ident, {}))
        entree["visible"] = _est_complete(entree)
        societes[ident] = entree

    visibles = sum(1 for s in societes.values() if s["visible"])
    dire(f"Comparables : {visibles} sociétés complètes sur {len(societes)}.")
    return {
        "source": chemin.name,
        "societes": societes,
    }


def _est_complete(s: dict) -> bool:
    """La société porte-t-elle toutes les valeurs dont la plateforme se sert ?

    Bêtas, dette et capitalisation alimentent le coût du capital ; chiffre
    d'affaires et résultat net alimentent l'affichage. L'EBITDA est hors critère,
    pour la raison exposée en tête de module.
    """
    if not s.get("pays") or not s.get("industrie"):
        return False
    if s.get("beta_1an") is None or s.get("beta_3ans") is None:
        return False
    if s.get("dette") is None or not s.get("capitalisation"):
        return False
    ca = s.get("ca") or {}
    rn = s.get("rn") or {}
    if any(ca.get(a) is None for a in EXERCICES):
        return False
    if any(rn.get(a) is None for a in EXERCICES[:3]):
        return False
    return True


def statistiques(membres: list) -> dict:
    """Bêtas et gearing médians d'un secteur, sur ses plus grosses valeurs.

    La médiane, et non la moyenne ni un rapport d'agrégats : une moyenne dit
    surtout ce que fait la plus grosse valeur du secteur, là où la médiane décrit
    la société typique — ce qu'on cherche pour un comparable.

    Mais la médiane de *toute* la population ne vaut pas mieux quand cette
    population est faite de micro-capitalisations qui ne s'échangent pas. Les 508
    banques nord-américaines complètes donnent un bêta médian à trois ans de
    0,32, alors que JPMorgan est à 0,90 et Bank of America à 0,98 : la médiane
    est portée par 374 établissements de moins de 760 millions d'euros, dont le
    bêta mesuré est proche de zéro faute d'être négocié. C'est exactement le
    défaut d'illiquidité qui écrase les bêtas africains, importé sur un marché où
    il est évitable.

    On retient donc les `ECHANTILLON_MAX` premières capitalisations du périmètre,
    ce qui rend 1,04 pour les banques nord-américaines et 0,92 pour les
    européennes. Un plancher de capitalisation absolu aurait le même effet sur
    ces deux marchés mais ne laisserait aucune société africaine : le classement
    par taille, lui, s'adapte à l'échelle de chaque marché et laisse l'Afrique
    inchangée, ses échantillons n'atteignant pas ce nombre.
    """
    def mediane(valeurs):
        return round(statistics.median(valeurs), 4) if valeurs else None

    retenus = sorted(membres, key=lambda s: -(s["capitalisation"] or 0))[:ECHANTILLON_MAX]
    return {
        # L'effectif du périmètre et l'échantillon qui produit les médianes sont
        # deux nombres différents dès que le secteur compte plus de vingt
        # sociétés. Les confondre ferait décrire une population et en chiffrer
        # une autre.
        "societes": len(membres),
        "retenues": len(retenus),
        "beta_1an": mediane([s["beta_1an"] for s in retenus]),
        "beta_3ans": mediane([s["beta_3ans"] for s in retenus]),
        "gearing": mediane([s["dette"] / s["capitalisation"] for s in retenus]),
        # Somme et non médiane : la capitalisation ne sert pas au calcul, elle
        # dit le poids de l'échantillon. Elle porte donc sur exactement les
        # sociétés qui produisent les médianes ci-dessus.
        "capitalisation": round(sum(s["capitalisation"] for s in retenus), 1),
    }


def secteurs_par_perimetre(societes: dict, classer) -> dict:
    """Statistiques de chaque secteur, à trois échelles emboîtées.

    Le CMPC se calcule sur la zone retenue. Mais l'univers est mince : sur les
    cent vingt-six couples zone x secteur, quarante pour cent ne comptent aucune
    société et un tiers n'en compte qu'une. Une médiane sur une seule société
    n'est pas une médiane, c'est cette société.

    On publie donc les trois échelles — zone, continent, univers — et la page
    prend la plus étroite qui atteigne le seuil, en écrivant laquelle a servi.
    Choisir à la construction figerait ce repli sans que l'utilisateur le voie.
    """
    paniers = {"univers": {}, "continents": {}, "zones": {}}
    for s in societes.values():
        if not s.get("visible"):
            continue
        nom = s["industrie"]
        continent, zone = classer(s["pays"])
        paniers["univers"].setdefault(nom, []).append(s)
        if continent:
            paniers["continents"].setdefault(nom, {}).setdefault(continent, []).append(s)
        if zone:
            paniers["zones"].setdefault(nom, {}).setdefault(zone, []).append(s)

    sortie = {}
    for nom, membres in sorted(paniers["univers"].items()):
        sortie[nom] = {
            "univers": statistiques(membres),
            "continents": {c: statistiques(m)
                           for c, m in sorted(paniers["continents"].get(nom, {}).items())},
            "zones": {z: statistiques(m)
                      for z, m in sorted(paniers["zones"].get(nom, {}).items())},
        }
    return sortie


def payload_page(univers: dict, classer) -> dict:
    """Les sociétés visibles, allégées pour être embarquées dans la page.

    C'est la seule source de comparables de la page : décomptes du cadrage,
    médianes du CMPC et cartes de l'onglet Sociétés en sortent toutes.

    Deux précautions qui ne se voient pas dans le résultat :

    - **Unités.** Produits, EBITDA et résultat net sont exportés en milliers
      (`BCEAO000`), la capitalisation en millions (`BCEAOM`). Tout est ramené
      ici en millions, faute de quoi la même page afficherait deux échelles
      sans le dire.

    Les descriptions d'activité ne sont pas embarquées : intégrales elles pesaient
    sept mégaoctets sur dix, et la page doit rester un fichier unique que le
    navigateur analyse d'un bloc. L'industrie fine de S&P, elle, est conservée :
    elle tient en quelques mots et suffit à situer l'activité.

    - **Zone.** Elle est calculée ici, par le `classer` qui produit aussi les
      effectifs du cadrage. La déduire côté navigateur à partir du pays
      supposerait que tout pays coté soit connu de Damodaran : le Malawi et la
      Gambie ne le sont pas, et leurs quatorze sociétés auraient été comptées
      sans jamais pouvoir être listées.
    """
    def millions(valeur):
        return None if valeur is None else round(valeur / 1000.0, 1)


    sortie = {}
    for ident, s in univers["societes"].items():
        if not s.get("visible"):
            continue
        continent, zone = classer(s["pays"])
        ca, ebitda, rn = s.get("ca") or {}, s.get("ebitda") or {}, s.get("rn") or {}
        sortie[ident] = {
            "nom": s["nom"],
            "ticker": s["ticker"],
            "place": s["place"],
            "pays": s["pays"],
            "continent": continent,
            "zone": zone,
            "industrie": s["industrie"],
            "activite": s.get("industrie_fine"),
            "beta_1an": s.get("beta_1an"),
            "beta_3ans": s.get("beta_3ans"),
            "capitalisation": round(s["capitalisation"], 1),
            "dette": round(s["dette"], 1),
            # Exercices en clair : le gabarit indexe les séries par année.
            "annees": [int(a[2:]) for a in reversed(EXERCICES)],
            "ca": {a[2:]: millions(ca.get(a)) for a in EXERCICES},
            "ebitda": {a[2:]: millions(ebitda.get(a)) for a in EXERCICES},
            "rn": {a[2:]: millions(rn.get(a)) for a in EXERCICES[:3]},
        }
    return sortie


def indexer_par_zone(univers: dict, classer) -> tuple[dict, list]:
    """Effectifs par zone, et par industrie au sein d'une zone.

    `classer(pays)` doit rendre (continent, zone). Les sociétés d'un pays
    qu'aucune zone ne réclame sont comptées à part plutôt qu'écartées : le build
    doit pouvoir les nommer.
    """
    index, orphelins = {}, []
    for s in univers["societes"].values():
        if not s.get("visible"):
            continue  # une société incomplète ne se compte pas sur la carte
        _continent, zone = classer(s["pays"])
        if zone is None:
            orphelins.append(s["pays"])
            continue
        place = s.get("place") or "hors cote"
        seau = index.setdefault(zone, {"total": 0, "places": {}, "industries": {}})
        seau["total"] += 1
        seau["places"][place] = seau["places"].get(place, 0) + 1
        detail = seau["industries"].setdefault(s["industrie"], {"n": 0, "places": {}})
        detail["n"] += 1
        detail["places"][place] = detail["places"].get(place, 0) + 1
    return index, sorted(set(orphelins))


def construire(racine: Path, progress=None) -> dict | None:
    """Point d'entrée du build : fusionne tous les exports déposés.

    Rend None si aucun n'est trouvé. Les sociétés sont appariées par `Entity ID`,
    identifiant S&P global : un même titre présent dans deux exports n'est donc
    compté qu'une fois, et la fusion reste sûre quand les périmètres se
    recouvrent — ce que fera tôt ou tard un export « Moyen-Orient » face à un
    export « Afrique » sur l'Égypte.
    """
    chemins = trouver_exports(racine)
    if not chemins:
        return None

    societes, sources, doublons = {}, [], 0
    for chemin in chemins:
        bloc = charger(chemin, progress=progress)
        doublons += len(set(bloc["societes"]) & set(societes))
        societes.update(bloc["societes"])
        sources.append(chemin.name)

    if progress:
        visibles = sum(1 for s in societes.values() if s["visible"])
        recouvrement = f", {doublons} en double" if doublons else ""
        progress(f"Comparables : {len(sources)} export(s) fusionné(s) — "
                 f"{visibles} sociétés complètes sur {len(societes)}{recouvrement}.")
    return {"source": ", ".join(sources), "societes": societes}
